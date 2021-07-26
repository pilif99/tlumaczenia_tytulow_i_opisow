from google_trans_new import google_translator
import openpyxl
import pyautogui

#klasa

class komorka_w_excelu:

    def __init__(self, rzad, kolumna):

        self.__rzad = rzad
        self.__kolumna = kolumna

    def get(self):
        return plik2.cell(self.__rzad, self.__kolumna).value

    def set(self, wartosc, jezyk):
        if jezyk == 'de':
            try:
                plik2.cell(self.__rzad, self.__kolumna + 7).value = wartosc
            except:
                print('wiersz: ' + str(wiersz))
        elif jezyk == 'es':
            try:
                plik2.cell(self.__rzad, self.__kolumna + 14).value = wartosc
            except:
                print('wiersz: ' + str(wiersz))
        elif jezyk == 'it':
            try:
                plik2.cell(self.__rzad, self.__kolumna + 21).value = wartosc
            except:
                print('wiersz: ' + str(wiersz))
    def get_rzad(self):
        return self.__rzad

    def get_kolumna(self):
        return self.__kolumna

#zmienne

plik = openpyxl.load_workbook(r"C:\Users\FLorenzLen\Documents\Seller\TEMPLATES\tytuły, opisy, bullet pointy.xlsx")
plik2 = plik.active
translator = google_translator()
wiersz = 100
slowa = ['Leather', 'Textile', 'Motorcycle', 'Cowhide', 'Goatskin', 'Jeans', 'For', 'Shirt', 'Black', 'Navy', 'Riding', 'Vintage', 'Rain', 'Sport', 'Top',
         'Thermoactive', 'Perforated']

#główna pętla

while plik2.cell(wiersz, 2).value != None or plik2.cell(wiersz, 3).value != None:
    
    """
    
    a = komorka_w_excelu(wiersz, 3)
    b = plik2.cell(a.get_rzad(), a.get_kolumna()).value
    x = b.split()
    j = 0
    while not x[j] in slowa:
        j += 1
    c = x.index(x[j])
    d = []
    for k in range(c):
        d.append(x[0])
        del x[0]
    translate_text = translator.translate(' '.join(x),lang_tgt='de')
    a.set(' '.join(d) + ' ' + translate_text, 'de')
    translate_text = translator.translate(' '.join(x),lang_tgt='es')
    a.set(' '.join(d) + ' ' + translate_text, 'es')
    translate_text = translator.translate(' '.join(x),lang_tgt='it')
    a.set(' '.join(d) + ' ' + translate_text, 'it')
    
    """

    for i in range(6):
        a = komorka_w_excelu(wiersz, i + 4)
        translate_text = translator.translate(a.get(),lang_tgt='de')
        a.set(translate_text, 'de')   
        translate_text = translator.translate(a.get(),lang_tgt='es')
        a.set(translate_text, 'es')   
        translate_text = translator.translate(a.get(),lang_tgt='it')
        a.set(translate_text, 'it')

    wiersz += 1
    print(wiersz)
    pyautogui.sleep(2)

#zapis pliku

plik.save(r"C:\Users\FLorenzLen\Documents\Seller\TEMPLATES\tytuły, opisy, bullet pointy.xlsx")